#!/usr/bin/env python3
"""SFGL 2019 reconstruction: repopulate broken raw slug tabs from ESPN JSON,
then compute SFGL values (money + round-leader bonuses) in Python for the DB.
Writes ONLY A1:I{n} in the raw tabs; L/M/N/O formulas and event tabs self-heal
in Google Sheets."""
import json, re, time, unicodedata, urllib.request
import openpyxl

WB_IN  = '/tmp/sfgl2019.xlsx'
WB_OUT = '/home/claude/hist/sfgl2019_reconstructed.xlsx'
JSON_OUT = '/home/claude/hist/sfgl2019_reconstructed_values.json'

# sheet event tab -> (slug tab, espn id)
EVENTS = {
 'Wells Fargo':          ('wells-fargo',                    '401056550'),
 'AT&T Byron Nelson':    ('at&t-byron-nelson',              '401056551'),
 'PGA CHAMPIONSHIP':     ('pga-championship',               '401056552'),
 'Charles Schwab':       ('charles-schwab-challenge',       '401056553'),
 'Memorial Tournament':  ('memorial-tournament',            '401056554'),
 'RBC Canadian':         ('rbc-canadian-open',              '401056555'),
 'US OPEN':              ('us-open-championship',           '401056556'),
 'Travelers':            ('the-travelers-championship',     '401056549'),
 'Rocket Mortgage':      ('rocket-mortgage-classic',        '401056557'),
 '3M Open':              ('3m-open',                        '401056558'),
 'John Deere':           ('john-deere-classic',             '401056548'),
 'OPEN CHAMPIONSHIP':    ('the-open-championship',          '401056547'),
 'WGC-FedEx St. Jude':   ('wgc-fedex-st-jude-invitational', '401056559'),
 'Wyndham':              ('wyndham-championship',           '401056545'),
 'Northern Trust':       ('the-northern-trust',             '401056544'),
 'BMW':                  ('bmw-championship',               '401056543'),
 'Tour Championship':    ('tour-championship',              '401056542'),  # FLAG: FedEx staggered/bonus
}
HEADER = ['PLAYER','POS','[TABLE]',None,None,None,'TOTAL\nSCORE','OFFICIAL\nMONEY','FEDEXCUP\nPOINTS']

def canonical(name):
    s = unicodedata.normalize('NFKD', name)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace('ø','o').replace('æ','ae').replace('đ','d').replace('ß','ss')
    s = re.sub(r"[.\-'’]", ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def fetch(eid, tries=3):
    url = f"https://site.api.espn.com/apis/site/v2/sports/golf/leaderboard?league=pga&event={eid}"
    for t in range(tries):
        try:
            req = urllib.request.Request(url, headers={'User-Agent':'Mozilla/5.0'})
            return json.load(urllib.request.urlopen(req, timeout=30))
        except Exception as e:
            if t == tries-1: raise
            time.sleep(2)

def parse_field(data):
    ev = data['events'][0]
    comp = ev['competitions'][0]
    rows = []
    for p in comp['competitors']:
        name = p['athlete']['displayName']
        st = p.get('status', {}) or {}
        sttype = (st.get('type', {}) or {}).get('name', '')
        pos_dv = ((st.get('position', {}) or {}).get('displayName')) or ''
        ls = p.get('linescores', []) or []
        rounds = [None, None, None, None]
        for l in ls:
            per = l.get('period')
            if isinstance(per, int) and 1 <= per <= 4 and isinstance(l.get('value'), (int, float)):
                rounds[per-1] = int(l['value'])
        earnings = p.get('earnings')
        cup = next((s.get('value') for s in p.get('statistics', []) if s.get('name') == 'cupPoints'), None)
        if sttype == 'STATUS_CUT':
            pos = 'CUT'
        elif sttype in ('STATUS_WITHDRAWN', 'STATUS_WD'):
            pos = 'W/D'
        elif sttype == 'STATUS_DISQUALIFIED':
            pos = 'DQ'
        else:
            pos = pos_dv
            # numeric like sheet ("1" -> 1)
            if isinstance(pos, str) and pos.isdigit(): pos = int(pos)
        total = sum(r for r in rounds if r is not None) if any(r is not None for r in rounds) else None
        money = earnings if isinstance(earnings, (int, float)) and earnings > 0 else None
        rows.append({'name': name, 'pos': pos, 'rounds': rounds, 'total': total,
                     'money': money, 'fedex': cup, 'amateur': bool(p.get('amateur'))})
    return ev['name'], rows

def read_bonus_constants(wbf, event_tab):
    """Parse the three round-leader bonus constants from the event tab formulas."""
    ws = wbf[event_tab]
    b = [None, None, None]
    for r in range(2, 15):
        for i, c in enumerate((2, 3, 4)):  # B,C,D
            v = ws.cell(r, c).value
            if isinstance(v, str) and 'min(indirect' in v:
                m = re.search(r'\)\)\),(\d+),""\)\)$', v.replace(' ', ''))
                if m and b[i] is None: b[i] = int(m.group(1))
        if all(b): break
    return b  # e.g. [10000,20000,30000] or [20000,40000,60000]

def read_rosters(wbv, event_tab):
    """Fixed block layout in event tab col A: team name at rows 2,9,16,23,30;
    five player rows follow each; 'Total' row closes the block. A blank cached
    Total cell must not bleed one block into the next, hence fixed offsets."""
    ws = wbv[event_tab]
    teams = {}
    for start in (2, 9, 16, 23, 30):
        tname = ws.cell(start, 1).value
        if not tname: continue
        players = []
        for r in range(start + 1, start + 6):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip() and v.strip() != 'Total':
                players.append(v.strip())
        teams[str(tname).strip()] = players
    return teams

def main():
    wbv = openpyxl.load_workbook(WB_IN, data_only=True)    # cached values (rosters/team names)
    wbf = openpyxl.load_workbook(WB_IN, data_only=False)   # formulas preserved; this one gets written
    out = {'season': 2019, 'events': [], 'flags': []}

    for event_tab, (slug, eid) in EVENTS.items():
        data = fetch(eid)
        espn_name, rows = parse_field(data)
        print(f"{event_tab:22s} <- ESPN {eid} '{espn_name}' field={len(rows)}")

        # name reconciliation: rewrite ESPN spelling to roster spelling where canonically equal
        rosters = read_rosters(wbv, event_tab)
        roster_names = [p for ps in rosters.values() for p in ps]
        canon_to_row = {}
        for row in rows: canon_to_row.setdefault(canonical(row['name']), row)
        unmatched = []
        for rn in roster_names:
            row = canon_to_row.get(canonical(rn))
            if row is None:
                unmatched.append(rn)      # not in field (didn't play) — fine
            elif row['name'] != rn:
                row['name'] = rn          # keep sheet vlookups working

        # write raw tab A1:I{n}
        ws = wbf[slug]
        for c, h in enumerate(HEADER, start=1): ws.cell(1, c).value = h
        for i, row in enumerate(rows, start=2):
            vals = [row['name'], row['pos'], *row['rounds'], row['total'], row['money'], row['fedex']]
            for c, v in enumerate(vals, start=1): ws.cell(i, c).value = v
        # clear any leftover junk below the field in A:I
        for r in range(len(rows) + 2, 201):
            for c in range(1, 10): ws.cell(r, c).value = None

        # python-side compute (mirror of sheet formulas)
        b1, b2, b3 = read_bonus_constants(wbf, event_tab)
        a1 = {r['name']: r['rounds'][0] for r in rows if r['rounds'][0] is not None and r['pos'] != 'W/D'}
        a2 = {r['name']: r['rounds'][0] + r['rounds'][1] for r in rows
              if r['rounds'][0] is not None and r['rounds'][1] is not None and r['pos'] != 'W/D'}
        a3 = {r['name']: r['rounds'][0] + r['rounds'][1] + r['rounds'][2] for r in rows
              if all(r['rounds'][k] is not None for k in (0, 1, 2)) and r['pos'] != 'W/D'}
        m1 = min(a1.values()) if a1 else None
        m2 = min(a2.values()) if a2 else None
        m3 = min(a3.values()) if a3 else None
        by_name = {r['name']: r for r in rows}

        ev_out = {'event': event_tab, 'slug': slug, 'espnId': eid, 'espnName': espn_name,
                  'bonuses': [b1, b2, b3], 'teams': {}}
        for team, players in rosters.items():
            entries, ttotal = [], 0
            for pn in players:
                r = by_name.get(pn)
                if r is None:
                    entries.append({'player': pn, 'played': False, 'value': 0})
                    continue
                money = r['money'] or 0
                bon = 0
                if b1 and a1.get(pn) == m1: bon += b1
                if b2 and a2.get(pn) == m2: bon += b2
                if b3 and a3.get(pn) == m3: bon += b3
                val = money + bon
                ttotal += val
                entries.append({'player': pn, 'played': True, 'pos': r['pos'],
                                'money': money, 'bonus': bon, 'value': val})
            ev_out['teams'][team] = {'players': entries, 'total': ttotal}
        out['events'].append(ev_out)
        if event_tab == 'Tour Championship':
            out['flags'].append('Tour Championship 2019: FedExCup staggered start + bonus payout; '
                                'ESPN earnings are the bonus money — REVIEW before trusting.')
        time.sleep(1)

    wbf.save(WB_OUT)
    json.dump(out, open(JSON_OUT, 'w'), indent=1)
    print("\nSaved:", WB_OUT, "and", JSON_OUT)

if __name__ == '__main__':
    main()
