#!/usr/bin/env python3
"""
SFGL historical extractor (staging pass).
Reads the 12 season-master workbooks (2015-2026) already downloaded to this dir
and produces a REVIEWABLE staged dataset:
  - sfgl_history_staged.json
  - sfgl_history_staged.xlsx  (PlayerSeasons / PlayerCareers / TeamSeasons / DataQuality)

Principle: DERIVE, don't trust. Summary/TOTAL cells are formula-fragile
(#REF!/#N/A), so every total is recomputed by summing the per-event columns.
Nothing here touches Firestore or prod. This is a review artifact only.
"""
import json, unicodedata, re
from collections import defaultdict
import openpyxl

YEARS = [str(y) for y in range(2015, 2027)]

# ---- confirmed franchise map: every label a franchise ever used -> manager ----
MANAGERS = {
    "TJ Crawforth":   ["detroit rock city", "drc"],
    "Josh Fano":      ["youll fing see", "you'll f***ing see", "yfs", "nubs", "twhooo",
                        "twistface", "gbd", "im baaaaccckk", "i'm baaaaccckk",
                        "im waaaaay baaaaccckk", "i'm waaaaay baaaaccckk", "baaaaccckk",
                        "iwb", "ss express", "hip happens"],
    "Michael Hershfield": ["dirty birdies", "dirty bird(ies)", "dirty birds", "db"],
    "Aaron Jensen":   ["air jordan", "the big cats back", "the big cat's back", "big cat",
                        "world #1", "world 1", "w#1", "w1"],
    "Dave Lutz":      ["mac12", "texas bulldozers", "tbdzr", "national champs", "natl champs",
                        "wtf point", "wtfp", "cttp", "nc", "pops llc", "pops, llc", "pops l l c"],
}
CURRENT_TEAM = {
    "TJ Crawforth": "Detroit Rock City", "Josh Fano": "Hip Happens",
    "Michael Hershfield": "Dirty Bird(ies)", "Aaron Jensen": "World #1",
    "Dave Lutz": "POPS, LLC",
}
UTILITY_TABS = {"draft", "rosters", "rostersold", "schedule", "schedule data",
                "transactions", "database_upload", "rosters_upload"}

def norm_label(s):
    if s is None: return ""
    s = str(s).lower().replace("(m)", " ")
    s = s.replace("*", " ")
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9#, ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

LABEL2MGR = {}
for mgr, labels in MANAGERS.items():
    for lab in labels:
        LABEL2MGR[norm_label(lab)] = mgr

def canonical_name(s):
    """Light approximation of the app's normalizeNordic()/canonicalName()."""
    if s is None: return ""
    s = str(s).strip()
    s = re.sub(r"^[\*\s]+", "", s)                    # strip leading keeper asterisks
    s = s.replace("ø", "o").replace("Ø", "o").replace("æ", "ae").replace("Æ", "ae")
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower().replace(".", "").replace(",", "").replace("'", "")
    s = s.replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def star_count(s):
    m = re.match(r"^(\*+)", str(s).strip())
    return len(m.group(1)) if m else 0

def display_name(s):
    return re.sub(r"^[\*\s]+", "", str(s)).strip()

def find_total_col(ws):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).strip().upper() == "TOTAL":
            return c
    return None

def header_map(ws):
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value not in (None, "")}

def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def extract_team_tab(ws, total_col, event_cols, starts_col):
    """Return (players, event_sums, cols_with_data). Keeps full roster even when
    earnings are unrecoverable (broken 'Loading...' imports), tagging those rows."""
    players = []
    event_sums = defaultdict(float)
    cols_with_data = set()
    r = 3
    blanks = 0
    while r <= ws.max_row and blanks < 2:
        name = ws.cell(r, 1).value
        if name in (None, ""):
            blanks += 1; r += 1; continue
        blanks = 0
        nm = display_name(name)
        low = norm_label(nm)
        # skip footer / team-name / non-player rows (but KEEP zero-data player rows)
        if low in LABEL2MGR or "total" in low or low in UTILITY_TABS or len(low) < 2:
            r += 1; continue
        nums = [(c, ws.cell(r, c).value) for c in event_cols if is_number(ws.cell(r, c).value)]
        earnings = sum(v for _, v in nums)
        starts = len(nums)
        for c, v in nums:
            event_sums[c] += v
            cols_with_data.add(c)
        stored_total = ws.cell(r, total_col).value
        stored_starts = ws.cell(r, starts_col).value if starts_col else None
        players.append({
            "name": nm, "canonical": canonical_name(nm), "stars": star_count(name),
            "_earnings": round(earnings, 2), "_starts": starts, "_hasData": starts > 0,
            "_stored_total_ok": is_number(stored_total),
            "_stored_starts": stored_starts if is_number(stored_starts) else None,
        })
        r += 1
    return players, event_sums, cols_with_data

def main():
    seasons = {}
    dq_rows = []
    unmatched = []
    for yr in YEARS:
        try:
            wb = openpyxl.load_workbook(f"sfgl{yr}.xlsx", data_only=True)
        except FileNotFoundError:
            continue
        year_teams = {}
        events_by_col = {}
        for ws in wb.worksheets:
            if ws.title.strip().lower() in UTILITY_TABS:
                continue
            total_col = find_total_col(ws)
            if not total_col:
                continue
            team_name = display_name(ws.cell(2, 1).value or "")
            mgr = LABEL2MGR.get(norm_label(team_name))
            event_cols = list(range(2, total_col))
            if len(event_cols) < 20:
                continue  # event mini-grid (5-team x 1 total), not a season player matrix
            # count player-name rows to distinguish team tabs from stray TOTAL tabs
            namerows = sum(1 for r in range(3, min(ws.max_row, 60) + 1)
                           if ws.cell(r, 1).value not in (None, "")
                           and "total" not in norm_label(ws.cell(r, 1).value))
            if not mgr:
                if namerows >= 8 and len(event_cols) >= 10 and norm_label(team_name):
                    unmatched.append({"year": yr, "tab": ws.title, "row2A": team_name})
                continue
            hdr = header_map(ws)
            starts_col = hdr.get("# of Starts")
            for c in event_cols:
                nm = ws.cell(1, c).value
                if nm not in (None, ""):
                    events_by_col.setdefault(c, str(nm).strip())
            players, event_sums, cols_with_data = extract_team_tab(ws, total_col, event_cols, starts_col)
            # completeness from whether the sheet's OWN totals survived (distinguishes
            # cancelled/blank events, e.g. COVID 2020, from genuinely broken imports)
            n = len(players) or 1
            intact_frac = sum(1 for p in players if p["_stored_total_ok"]) / n
            data_frac = sum(1 for p in players if p["_hasData"]) / n
            if yr == "2026":
                completeness = "future/empty"
            elif intact_frac >= 0.45:
                completeness = "complete"
            elif data_frac < 0.15:
                completeness = "rosters-only"
            else:
                completeness = "partial"
            usable = completeness in ("complete", "partial")
            for p in players:
                if usable and p["_hasData"]:
                    p["earnings"] = p["_earnings"]; p["starts"] = p["_starts"]
                    p["perStart"] = round(p["_earnings"] / p["_starts"], 2) if p["_starts"] else 0
                else:
                    p["earnings"] = None; p["starts"] = None; p["perStart"] = None
            season_earnings = round(sum(p["_earnings"] for p in players), 2) if usable else None
            broken = sum(1 for p in players if not p["_stored_total_ok"])
            scheck = [p for p in players if p["_stored_starts"] is not None]
            starts_ok = sum(1 for p in scheck if p["_stored_starts"] == p["_starts"])
            dq_rows.append({
                "year": yr, "manager": mgr, "team_tab": team_name, "completeness": completeness,
                "players": len(players), "events_with_data": f"{len(cols_with_data)}/{len(event_cols)}",
                "starts_crosscheck": f"{starts_ok}/{len(scheck)}" if scheck else "n/a",
            })
            year_teams[mgr] = {
                "teamThatYear": team_name, "seasonEarnings": season_earnings,
                "completeness": completeness, "players": players,
                "eventSums": {str(k): round(v, 2) for k, v in event_sums.items()},
            }
        # rank + naive event wins (only for years with usable earnings)
        rankable = [(mgr, t) for mgr, t in year_teams.items() if t["seasonEarnings"] is not None]
        for mgr, t in year_teams.items():
            t["rank"] = None
        for i, (mgr, t) in enumerate(sorted(rankable, key=lambda kv: -kv[1]["seasonEarnings"]), 1):
            t["rank"] = i
        wins = defaultdict(int)
        for c in events_by_col:
            best = None; bestmgr = None
            for mgr, t in year_teams.items():
                v = t["eventSums"].get(str(c), 0)
                if v and (best is None or v > best):
                    best, bestmgr = v, mgr
            if bestmgr:
                wins[bestmgr] += 1
        for mgr, t in year_teams.items():
            t["naiveEventWins"] = wins.get(mgr, 0)
        seasons[yr] = {"events": [events_by_col[c] for c in sorted(events_by_col)],
                       "teams": year_teams}

    # ---- career rollups ----
    careers = {}
    for yr, sd in seasons.items():
        for mgr, t in sd["teams"].items():
            complete = t["completeness"] == "complete"
            for p in t["players"]:
                key = p["canonical"]
                if not key: continue
                c = careers.setdefault(key, {
                    "displayName": p["name"], "totalEarnings": 0.0, "totalStarts": 0,
                    "seasons": set(), "partialSeasons": set(), "teams": set(),
                    "best": {"year": None, "earnings": 0}})
                c["seasons"].add(yr)
                c["teams"].add(mgr)
                if complete and p["earnings"] is not None:
                    c["totalEarnings"] += p["earnings"]
                    c["totalStarts"] += p["starts"]
                    if p["earnings"] > c["best"]["earnings"]:
                        c["best"] = {"year": yr, "earnings": p["earnings"]}
                else:
                    c["partialSeasons"].add(yr)
    for c in careers.values():
        c["totalEarnings"] = round(c["totalEarnings"], 2)
        c["seasonsPlayed"] = len(c["seasons"])
        c["seasons"] = sorted(c["seasons"])
        c["partialSeasons"] = sorted(c["partialSeasons"])
        c["teams"] = sorted(c["teams"])
        c["perStart"] = round(c["totalEarnings"] / c["totalStarts"], 2) if c["totalStarts"] else 0

    # strip private fields from json
    out_seasons = {}
    for yr, sd in seasons.items():
        teams = {}
        for mgr, t in sd["teams"].items():
            teams[mgr] = {k: v for k, v in t.items() if k != "eventSums"}
            for p in teams[mgr]["players"]:
                for pk in ("_stored_total_ok", "_stored_starts", "_earnings", "_starts", "_hasData"):
                    p.pop(pk, None)
        out_seasons[yr] = {"events": sd["events"], "teams": teams}
    json.dump({"seasons": out_seasons, "careers": careers},
              open("/mnt/user-data/outputs/sfgl-history/sfgl_history_staged.json", "w"), indent=1)

    # ---- xlsx ----
    from openpyxl import Workbook
    out = Workbook(); out.remove(out.active)
    ps = out.create_sheet("PlayerSeasons")
    ps.append(["Year", "Manager", "Team That Year", "Player", "SFGL Earnings", "Starts", "$/Start", "Keeper *", "Completeness"])
    for yr in YEARS:
        if yr not in seasons: continue
        for mgr, t in sorted(seasons[yr]["teams"].items()):
            for p in sorted(t["players"], key=lambda x: -(x["earnings"] or -1)):
                ps.append([yr, mgr, t["teamThatYear"], p["name"], p["earnings"], p["starts"],
                           p["perStart"], "*"*p["stars"], t["completeness"]])
    pc = out.create_sheet("PlayerCareers")
    pc.append(["Player", "Complete Seasons", "Seasons Played", "Teams", "Total SFGL Earnings (complete yrs)",
               "Total Starts", "$/Start", "Best Year", "Best-Year Earnings", "Partial/Roster-only Seasons"])
    for key, c in sorted(careers.items(), key=lambda kv: -kv[1]["totalEarnings"]):
        pc.append([c["displayName"], ",".join(c["seasons"]), c["seasonsPlayed"], ",".join(c["teams"]),
                   c["totalEarnings"], c["totalStarts"], c["perStart"], c["best"]["year"],
                   c["best"]["earnings"], ",".join(c["partialSeasons"])])
    ts = out.create_sheet("TeamSeasons")
    ts.append(["Year", "Manager", "Current Team", "Team That Year", "Season Earnings (derived)", "Rank", "Naive Event Wins", "Completeness"])
    for yr in YEARS:
        if yr not in seasons: continue
        for mgr, t in sorted(seasons[yr]["teams"].items(), key=lambda kv: (kv[1]["rank"] is None, kv[1]["rank"] or 99)):
            ts.append([yr, mgr, CURRENT_TEAM[mgr], t["teamThatYear"], t["seasonEarnings"], t["rank"], t["naiveEventWins"], t["completeness"]])
    dq = out.create_sheet("DataQuality")
    dq.append(["Year", "Manager", "Team Tab", "Completeness", "Players Found", "Events With Data", "Starts Cross-check"])
    for r in dq_rows:
        dq.append([r["year"], r["manager"], r["team_tab"], r["completeness"], r["players"], r["events_with_data"], r["starts_crosscheck"]])
    out.save("/mnt/user-data/outputs/sfgl-history/sfgl_history_staged.xlsx")

    # ---- console summary ----
    print("SEASONS:", ", ".join(y for y in YEARS if y in seasons))
    print(f"CAREER PLAYERS: {len(careers)}")
    for yr in YEARS:
        if yr not in seasons: continue
        comp = defaultdict(int)
        for t in seasons[yr]["teams"].values():
            comp[t["completeness"]] += 1
        ps_n = sum(1 for t in seasons[yr]["teams"].values() for p in t["players"])
        print(f"  {yr}: teams={len(seasons[yr]['teams'])}/5  players={ps_n}  {dict(comp)}")
    if unmatched:
        print("UNMATCHED team-looking tabs (verify manager):")
        for u in unmatched:
            print(f"   {u['year']}: tab={u['tab']!r} name={u['row2A']!r}")
    for yr in seasons:
        missing = set(MANAGERS) - set(seasons[yr]["teams"])
        if missing:
            print(f"  !! {yr} missing franchises: {missing}")

if __name__ == "__main__":
    import os
    os.makedirs("/mnt/user-data/outputs/sfgl-history", exist_ok=True)
    main()
