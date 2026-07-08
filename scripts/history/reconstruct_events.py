#!/usr/bin/env python3
"""
SFGL 2018/2019 event reconstruction pipeline.

Purpose: the 2018 (37 events) and 2019 (18 events) season masters have raw
leaderboard tabs that were IMPORTHTML live-pulls from pgatour.com. That step's
"paste as values" was skipped, the URLs later died, and the tabs now read
"Loading..."/#REF! -> all downstream scoring is broken.

This script repopulates each broken raw slug tab from ESPN's public JSON API
(full field: rounds + earnings), after which the sheet's own formulas recompute.
It ALSO computes each started player's SFGL value directly (money + round-leader
bonuses) so the historical DB doesn't depend on a spreadsheet recalc.

REQUIRES: network egress to site.api.espn.com (add via Settings > Capabilities >
Allow network egress). Verify with the smoke test in main() -- it re-proves the
2019 US Open (Woodland 2,350,000 / Rose 601,872) end-to-end before any batch run.

Raw-tab column layout (verified against intact 2019 front-half tabs, e.g.
'sony-open-in-hawaii'):
  A=player  B=pos  C=R1 D=R2 E=R3 F=R4  G=total  H=OFFICIAL MONEY  I=fedex
  L=player(helper)  M=After1(=R1)  N=After2(=R1+R2)  O=After3(=R1+R2+R3)
Event-tab formulas read raw cols: 2(pos), 8(money), 13/14/15 (After1/2/3).
Event-tab bonus constants (B/C/D) are read from the event tab itself (majors
20/40/60k, regular 10/20/30k) -- do NOT hardcode; parse them.

DECISIONS (confirmed with Aaron): missed cut => money $0. Round-leader bonus
fires on ties (formula uses == min()). WGC Match Play: money only, no rounds/
bonuses. 2019 Tour Championship: FedEx staggered start + bonus money -- handle
as a flagged special case (review before trusting).
"""
import json, re, subprocess, unicodedata
from collections import defaultdict
import openpyxl

# NOTE (2026-07-08): ESPN moved the leaderboard endpoint. The old
# `.../golf/pga/leaderboard?event={id}` path now 404s (it routes to
# leagues/all/events/{id}, which no longer resolves). The working form is
# `.../golf/leaderboard?league=pga&event={id}` -- same {"events":[...]} shape,
# carries per-competitor `earnings` + `linescores` + cut `status`.
ESPN_LB  = "https://site.api.espn.com/apis/site/v2/sports/golf/leaderboard?league=pga&event={eid}"
ESPN_SCB = "https://site.api.espn.com/apis/site/v2/sports/golf/pga/scoreboard?dates={year}"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"

def _curl_json(url):
    out = subprocess.run(["curl", "-sS", "-m", "30", "-A", UA, url],
                         capture_output=True, text=True)
    body = out.stdout
    if out.returncode != 0 or not body.strip():
        raise SystemExit(f"ESPN fetch failed ({out.returncode}). "
                         f"Is site.api.espn.com allowlisted? stderr: {out.stderr[:200]}")
    data = json.loads(body)
    # ESPN error envelope: {"code":404,"message":...} -- fail loud, don't return junk.
    if isinstance(data, dict) and data.get("code") and "message" in data:
        raise SystemExit(f"ESPN returned error {data['code']}: {data['message'][:180]}\n"
                         f"URL: {url}\n(If this is a path/param change, update ESPN_LB.)")
    return data

def fetch_espn(eid):
    """Return parsed ESPN leaderboard JSON for a tournament id (via curl in-sandbox)."""
    return _curl_json(ESPN_LB.format(eid=eid))

def fetch_schedule(year):
    """Return {canonical(event name): espn_event_id} for a PGA season.
    Uses the scoreboard endpoint (dates={year}), which lists the full season
    with ids -- more reliable than scraping the static schedule HTML."""
    d = _curl_json(ESPN_SCB.format(year=year))
    out = {}
    for e in d.get("events", []):
        eid, nm = e.get("id"), e.get("name")
        if eid and nm:
            out[canonical(nm)] = eid
    return out

def parse_field(espn_json):
    """Extract [{name, pos, rounds:[r1..], total, earnings, made_cut}] from ESPN JSON.
    NOTE: verify these key paths against a live dump on first run (structure below
    matches ESPN site v2 golf leaderboard as of prior seasons)."""
    ev = espn_json["events"][0]
    comp = ev["competitions"][0]
    field = []
    for c in comp["competitors"]:
        ath = c.get("athlete", {})
        name = ath.get("displayName") or ath.get("shortName") or ""
        # rounds from linescores
        rounds = []
        for ls in c.get("linescores", []):
            v = ls.get("value")
            if isinstance(v, (int, float)):
                rounds.append(int(v))
        # earnings: try common locations
        earnings = 0
        for stat in c.get("statistics", []):
            if stat.get("name") in ("earnings", "prizeMoney", "money"):
                try: earnings = float(str(stat.get("value")).replace("$", "").replace(",", ""))
                except Exception: pass
        if not earnings and c.get("earnings"):
            try: earnings = float(str(c["earnings"]).replace("$", "").replace(",", ""))
            except Exception: pass
        status = (c.get("status") or {})
        pos = (status.get("position") or {}).get("displayName") or status.get("displayValue") or ""
        made_cut = "cut" not in str(status.get("displayValue", "")).lower() and len(rounds) >= 3
        if not made_cut:
            earnings = 0.0  # confirmed: MC => $0
        field.append({"name": name, "pos": pos, "rounds": rounds,
                      "total": sum(rounds) if rounds else None,
                      "earnings": earnings, "made_cut": made_cut})
    return field

def canonical(s):
    s = re.sub(r"^[\*\s]+", "", str(s or ""))
    s = s.replace("ø","o").replace("æ","ae")
    s = unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode()
    return re.sub(r"\s+"," ", s.lower().replace(".","").replace(",","").replace("'","").replace("-"," ")).strip()

def populate_raw(ws, field):
    """Write the full ESPN field into a raw slug tab in the A..O layout."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, 18):
            ws.cell(r, c).value = None
    hdr = ['PLAYER','POS','R1','R2','R3','R4','TOTAL','OFFICIAL MONEY','FEDEX',
           None,None,'PLAYER','After 1','After 2','After 3']
    for c, h in enumerate(hdr, 1):
        ws.cell(1, c).value = h
    for i, p in enumerate(sorted(field, key=lambda x: (x["total"] is None, x["total"] or 9999)), start=2):
        r = (p["rounds"] + [None]*4)[:4]
        a1 = r[0]
        a2 = (r[0]+r[1]) if (r[0] and r[1]) else None
        a3 = (r[0]+r[1]+r[2]) if (r[0] and r[1] and r[2]) else None
        row = [p["name"], i-1, r[0], r[1], r[2], r[3], p["total"], p["earnings"], None,
               None, None, p["name"], a1, a2, a3]
        for c, v in enumerate(row, 1):
            ws.cell(i, c).value = v

def event_bonus_constants(event_ws):
    """Read the B/C/D award constants from an event tab (major 20/40/60k vs 10/20/30k)."""
    consts = []
    for col in (2, 3, 4):
        f = event_ws.cell(3, col).value or ""
        m = re.findall(r"\),(\d+),", f)
        consts.append(int(m[0]) if m else 0)
    return consts  # [B,C,D]

def sfgl_values(field, rostered_names, consts):
    """Compute SFGL value = money + round-leader bonuses for each rostered player."""
    byc = {canonical(p["name"]): p for p in field}
    min1 = min((p["rounds"][0] for p in field if len(p["rounds"])>=1), default=None)
    min2 = min((sum(p["rounds"][:2]) for p in field if len(p["rounds"])>=2), default=None)
    min3 = min((sum(p["rounds"][:3]) for p in field if len(p["rounds"])>=3), default=None)
    out = {}
    for nm in rostered_names:
        p = byc.get(canonical(nm))
        if not p:
            out[nm] = {"value": None, "note": "no ESPN match"}; continue
        v = p["earnings"]; bonus = 0
        R = p["rounds"]
        if len(R)>=1 and R[0]==min1: bonus += consts[0]
        if len(R)>=2 and sum(R[:2])==min2: bonus += consts[1]
        if len(R)>=3 and sum(R[:3])==min3: bonus += consts[2]
        out[nm] = {"value": v + bonus, "money": v, "bonus": bonus}
    return out

# 2019 event tab -> ESPN tournament id (CONFIRMED ones; auto-derive the rest with
# fetch_schedule(2019), then match canonical(event-tab name) -> id).
ESPN_IDS_2019_CONFIRMED = {
    "US OPEN": 401056556, "PGA CHAMPIONSHIP": 401056552, "OPEN CHAMPIONSHIP": 401056547,
    "BMW": 401056543, "Tour Championship": 401056542, "WGC Matchplay": 401056524,
}

def main():
    print("Smoke test ESPN access + reconstruction math (2019 US Open)...")
    d = fetch_espn(401056556)  # 2019 US Open (major: 20/40/60k round-leader bonuses)
    field = parse_field(d)
    print(f"  parsed {len(field)} competitors")
    vals = sfgl_values(field, ["Gary Woodland", "Justin Rose"], consts=[20000, 40000, 60000])
    expect = {"Gary Woodland": 2350000, "Justin Rose": 601872}
    ok = True
    for nm, exp in expect.items():
        got = vals[nm]["value"]
        flag = "OK" if got == exp else "MISMATCH"
        if got != exp: ok = False
        print(f"  {nm:16s} money={vals[nm]['money']:>10,.0f} +bonus={vals[nm]['bonus']:>6,} "
              f"=> SFGL={got:>10,.0f}  expect {exp:,}  [{flag}]")
    if not ok:
        raise SystemExit("Reconstruction math changed -- dump a competitor JSON and re-check "
                         "parse_field key paths before trusting any batch run:\n"
                         "  json.dump(d['events'][0]['competitions'][0]['competitors'][0], "
                         "open('one.json','w'), indent=1)")
    print("  PASS -- endpoint + parse + bonus logic all verified.")

if __name__ == "__main__":
    main()
