# Extract the SFGL 2026 Google Sheet into the JSON the audit reads.
#
# Usage:
#   pip install openpyxl
#   # File > Download > Microsoft Excel (.xlsx) from the sheet, then:
#   python3 scripts/extract-sheet-2026.py "SFGL 2026.xlsx"
#   python3 scripts/extract-sheet-2026.py in.xlsx --out scripts/fixtures/sheet-state-2026.json
#
# READ-ONLY on the sheet side: it opens a downloaded workbook and writes one
# JSON file. It never touches Firestore and needs no credentials.
#
# WHY THIS IS A SCRIPT AND NOT A ONE-OFF
# --------------------------------------
# The fixture it produces is the sheet half of every sheet-vs-app comparison,
# so "where did this number come from" has to stay answerable months later. A
# hand-built fixture answers that with "someone read the sheet once".
#
# WHAT IT KNOWS ABOUT THE WORKBOOK
# --------------------------------
# Four kinds of tab, each with a fixed shape:
#
#   Rosters        current roster per team, the two mulligans, the overall
#                  standings, and the per-swing earnings/transaction table.
#   Transactions   one column group per event (Claim/Add, Drop, Fee), one row
#                  band per team, up to four transactions deep.
#   <team> tabs    a player x tournament earnings grid. A BLANK cell means the
#                  player was not started; a 0 means he was started and earned
#                  nothing. That distinction is the whole lineup signal, so
#                  blanks are never coerced to zero.
#   <event> tabs   per-team blocks of starter rows ending in a Total row, plus
#                  the full field's official money off to the right.
#
# The three money views (event tabs, team tabs, Rosters standings) are all
# extracted rather than just the one, because they are independently maintained
# inside the same workbook and disagreeing with each other is itself a finding.

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

TEAMS = ['Detroit Rock City', 'Hip Happens', 'Dirty Bird(ies)', 'World #1', 'POPS, LLC']

# The Transactions tab still bands the second team under its old name. Team
# identity is by display name everywhere else in this file, so the alias is
# resolved on the way in rather than leaking a sixth team into the output.
TEAM_ALIASES = {'IWB': 'Hip Happens'}

# Tabs that are not tournaments. Everything else with the event-tab shape is.
NON_EVENT_TABS = {
    'RostersOLD', 'DRAFT', 'Rosters', 'Schedule', 'Schedule Data',
    'Transactions', 'APPTourneyResults', 'tour-championship',
    *TEAMS,
}


def num(v):
    """Cell → float, or None. Strings that are not numbers (#REF!, '--') → None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('$', '').replace(',', '')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def text(v):
    return None if v is None else str(v).strip() or None


def cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def extract_rosters(wb):
    """Rosters tab: rosters, keeper marks, mulligans, standings, swing table."""
    ws = wb['Rosters']
    out = {'rosters': {}, 'mulligans': {}, 'standings': [], 'swings': {}, 'summary': {}}

    # Row 1 carries a team name every third column starting at C, with the
    # roster's total starts budget beside it.
    team_cols = {}
    for col in range(1, ws.max_column + 1):
        name = text(cell(ws, 1, col))
        if name in TEAMS:
            team_cols[name] = col

    for name, col in team_cols.items():
        players = []
        for row in range(2, 15):          # 14 roster slots
            player = text(cell(ws, row, col))
            if not player:
                continue
            players.append({
                'player': player,
                # '***' / '**' / '*' — keeper tier, carried through because the
                # app stores it nowhere and only the sheet can answer for it.
                'keeper': text(cell(ws, row, col - 1)),
                'startsAllowed': num(cell(ws, row, col + 1)),
            })
        out['rosters'][name] = players
        out['startsBudget'] = out.get('startsBudget', {})
        out['startsBudget'][name] = num(cell(ws, 1, col + 1))

        # Rows 15-18: the signature mulligan (label, event / player) and the
        # non-signature one below it. A row with the event but no player under
        # it means the mulligan is still unused.
        out['mulligans'][name] = {
            'signature': {'event': text(cell(ws, 15, col)), 'player': text(cell(ws, 16, col))},
            'nonSignature': {'event': text(cell(ws, 17, col)), 'player': text(cell(ws, 18, col))},
        }

    # Standings, rows 22-26. Column G is an UNLABELLED helper column holding the
    # same five totals in tab order, so it does not line up with the sorted team
    # names beside it; column I is the one under the "Total" header. Reading G
    # would silently attribute every team the previous team's money.
    for row in range(22, 27):
        team = text(cell(ws, row, 6))
        if not team:
            continue
        out['standings'].append({
            'pos': num(cell(ws, row, 5)),
            'team': team,
            'total': num(cell(ws, row, 9)),
            'behindLeader': num(cell(ws, row, 12)),
        })

    # Swing table, rows 30-34: T$ (season fees), then per swing a transaction
    # count, the swing's earnings, and the fees collected in it.
    swing_cols = [('West Coast Swing', 5, 6, 7), ('Spring Swing', 8, 9, 10),
                  ('Summer Swing', 11, 12, 13), ('Fall Finish', 14, 15, 16)]
    for row in range(30, 35):
        team = text(cell(ws, row, 3))
        if not team:
            continue
        entry = {'seasonFees': num(cell(ws, row, 4)), 'bySwing': {}}
        for swing, txcol, earncol, feecol in swing_cols:
            entry['bySwing'][swing] = {
                'transactions': num(cell(ws, row, txcol)),
                'earnings': num(cell(ws, row, earncol)),
                'fees': num(cell(ws, row, feecol)),
            }
        out['swings'][team] = entry

    # Row 35 names each swing's winner and the pot they took.
    out['swingWinners'] = {}
    for swing, _tx, namecol, potcol in swing_cols:
        out['swingWinners'][swing] = {
            'team': text(cell(ws, 35, namecol)),
            'pot': num(cell(ws, 35, potcol)),
        }

    # Rows 38-42: wins, no-start winners, missed cuts, best week.
    for row in range(38, 43):
        team = text(cell(ws, row, 3))
        if not team:
            continue
        out['summary'][team] = {
            'wins': num(cell(ws, row, 6)),
            'noStartWinners': num(cell(ws, row, 9)),
            'missedCuts': num(cell(ws, row, 12)),
            'bestWeek': num(cell(ws, row, 15)),
        }

    return out


def extract_schedule(wb):
    """Schedule tab: the event list, its swing, and whether it counts for SFGL."""
    ws = wb['Schedule']
    events = []
    for row in range(3, ws.max_row + 1):
        name = text(cell(ws, row, 6))
        if not name:
            continue
        events.append({
            'event': num(cell(ws, row, 1)),
            'name': name,
            'swing': text(cell(ws, row, 2)),
            'sfgl': cell(ws, row, 3) is True or text(cell(ws, row, 3)) == 'TRUE',
            'signature': cell(ws, row, 4) is True or text(cell(ws, row, 4)) == 'TRUE',
            'dates': text(cell(ws, row, 5)),
            'purse': num(cell(ws, row, 9)),
        })
    return events


def extract_transactions(wb):
    """Transactions tab → a flat list of {team, event, added, dropped, fee}."""
    ws = wb['Transactions']

    # Row 1 labels each three-column group with its event, but only over the
    # first column of the group; the "Total Period"/"Total <swing>" columns are
    # summaries, not events, and must not become transaction slots.
    groups = []
    for col in range(1, ws.max_column + 1):
        label = text(cell(ws, 1, col))
        if not label or label.startswith('Total'):
            continue
        if text(cell(ws, 2, col)) in ('Claim/Add', 'Add'):
            groups.append((label, col))

    # Each team owns a band: a name row, then up to four transaction rows.
    bands = {}
    for row in range(1, ws.max_row + 1):
        name = text(cell(ws, row, 1))
        resolved = TEAM_ALIASES.get(name, name)
        if resolved in TEAMS:
            bands[resolved] = row

    rows = []
    for team, header_row in bands.items():
        for offset in range(1, 5):
            row = header_row + offset
            for event, col in groups:
                added = text(cell(ws, row, col))
                dropped = text(cell(ws, row, col + 1))
                fee = num(cell(ws, row, col + 2))
                if not added and not dropped and fee is None:
                    continue
                rows.append({
                    'team': team,
                    'event': event,
                    'added': added,
                    'dropped': dropped,
                    'fee': fee,
                    'slot': offset,
                })
    return rows


def extract_team_tab(ws):
    """A team tab: row 1 events, row 2 team totals, rows 3+ the player grid.

    Returns (totalsByEvent, playersByEvent) where playersByEvent[event] is the
    list of players holding a NON-BLANK cell for that event — the sheet's record
    of who was started, with 0 meaning started-and-earned-nothing.
    """
    events = {}
    for col in range(2, ws.max_column + 1):
        name = text(cell(ws, 1, col))
        if name and name not in ('TOTAL',):
            events[col] = name

    totals = {}
    for col, name in events.items():
        totals[name] = num(cell(ws, 2, col))

    started = {name: [] for name in events.values()}
    season = {}
    for row in range(3, 20):
        player = text(cell(ws, row, 1))
        if not player:
            continue
        season[player] = {
            'total': num(cell(ws, row, 34)),
            'starts': num(cell(ws, row, 35)),
        }
        for col, name in events.items():
            raw = cell(ws, row, col)
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                continue  # blank = not started. Never coerce to 0.
            started[name].append({'player': player, 'earnings': num(raw) or 0.0})

    return totals, started, season


def extract_event_tab(ws):
    """An event tab: per-team starter blocks, plus the full field to the right.

    Column layout inside a block:
      A player   B/C/D round-1/2/3 leader bonus   E place   F earnings
      G total    H the lineup as submitted
    """
    teams = {}
    row = 1
    last = ws.max_row
    while row <= last:
        name = text(cell(ws, row, 1))
        resolved = TEAM_ALIASES.get(name, name)
        if resolved in TEAMS and text(cell(ws, row, 2)) == 'Round 1':
            players = []
            submitted = []
            empty_slots = 0
            r = row + 1
            # Scan to the block's Total row. A blank column A does NOT end the
            # block: unused lineup slots are left empty in the middle of one,
            # and stopping at the first blank truncated 20 of the 31 events —
            # silently, because a short block still totals to something. The
            # Total row is the only terminator.
            while r <= last and text(cell(ws, r, 1)) != 'Total':
                label = text(cell(ws, r, 1))
                # Column H is the manager's submitted five, typed separately
                # from the scored rows. Where the two disagree the sheet is
                # arguing with itself, so both are carried — including the case
                # where H names a starter that has no earnings row at all.
                sub = text(cell(ws, r, 8))
                if sub:
                    submitted.append(sub)
                if label is None:
                    empty_slots += 1
                    r += 1
                    continue
                bonus = sum(num(cell(ws, r, c)) or 0.0 for c in (2, 3, 4))
                players.append({
                    'player': label,
                    'place': text(cell(ws, r, 5)),
                    'earnings': num(cell(ws, r, 6)) or 0.0,
                    'bonus': bonus,
                    'total': num(cell(ws, r, 7)) or 0.0,
                })
                r += 1
            total = num(cell(ws, r, 7)) if r <= last and text(cell(ws, r, 1)) == 'Total' else None
            teams[resolved] = {
                'total': total if total is not None else 0.0,
                'earnings': sum(p['earnings'] for p in players),
                'bonuses': sum(p['bonus'] for p in players),
                'earners': sum(1 for p in players if p['earnings'] > 0),
                'starters': len(players),
                # Lineup slots left blank in this block. A starter named in
                # column H with no row beside him lands here, so a lineup that
                # is short on the sheet is visible as short rather than as a
                # team that just happened to start four.
                'emptySlots': empty_slots,
                'players': players,
                'submitted': submitted,
            }
            row = r + 1
            continue
        row += 1

    # The field, columns J onward: name, position, four rounds, score, to par,
    # official money, then the after-round-N running totals the round-leader
    # bonuses are read off.
    field = []
    for r in range(3, ws.max_row + 1):
        player = text(cell(ws, r, 10))
        if not player:
            continue
        field.append({
            'player': player,
            'pos': text(cell(ws, r, 11)),
            'money': num(cell(ws, r, 18)) or 0.0,
            'after1': num(cell(ws, r, 20)),
            'after2': num(cell(ws, r, 21)),
            'after3': num(cell(ws, r, 22)),
        })

    return teams, field


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    if not args:
        sys.exit('Usage: python3 scripts/extract-sheet-2026.py <SFGL 2026.xlsx> [--out path.json]')
    src = Path(args[0])

    out_path = Path(__file__).parent / 'fixtures' / 'sheet-state-2026.json'
    if '--out' in sys.argv:
        out_path = Path(sys.argv[sys.argv.index('--out') + 1])

    wb = openpyxl.load_workbook(src, data_only=True)

    rosters = extract_rosters(wb)
    out = {
        'season': 2026,
        'source': f"Google Sheet 'SFGL 2026' via {src.name} (ESPN official money)",
        'rosters': rosters['rosters'],
        'startsBudget': rosters.get('startsBudget', {}),
        'mulligans': rosters['mulligans'],
        'standings': rosters['standings'],
        'swings': rosters['swings'],
        'swingWinners': rosters['swingWinners'],
        'summary': rosters['summary'],
        'schedule': extract_schedule(wb),
        'transactions': extract_transactions(wb),
        'teamTabs': {},
        'tournaments': {},
        'field': {},
        'seasonTotals': {},
    }

    for team in TEAMS:
        totals, started, season = extract_team_tab(wb[team])
        out['teamTabs'][team] = {'byEvent': totals, 'started': started, 'season': season}

    for ws in wb.worksheets:
        if ws.title in NON_EVENT_TABS or re.fullmatch(r'Sheet\d+', ws.title):
            continue
        teams, field = extract_event_tab(ws)
        if not teams:
            continue
        out['tournaments'][ws.title] = {'teams': teams}
        out['field'][ws.title] = field

    # Season totals from the EVENT tabs, which is the level the earnings diff
    # descends to. The Rosters standings and the team tabs are carried
    # separately above rather than folded in here — three sources, three
    # numbers, and flattening them would hide which one is wrong.
    for tourney in out['tournaments'].values():
        for team, block in tourney['teams'].items():
            out['seasonTotals'][team] = out['seasonTotals'].get(team, 0.0) + (block['total'] or 0.0)

    out_path.write_text(json.dumps(out, indent=1, ensure_ascii=False) + '\n', encoding='utf-8')

    print(f'Wrote {out_path}')
    print(f'  {len(out["tournaments"])} event tabs, {len(out["transactions"])} transactions')
    for team, total in sorted(out['seasonTotals'].items(), key=lambda kv: -kv[1]):
        print(f'  {team:<22} ${total:,.0f}')


if __name__ == '__main__':
    main()
